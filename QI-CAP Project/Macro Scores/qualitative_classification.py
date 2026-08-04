#qualitative classification of regimes for each pillar
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_styled_macro_regimes(input_xlsx_path, output_xlsx_path):
    # =========================================================================
    # STEP 1: LOAD NATIVE EXCEL & GENERATE QUALITATIVE COLUMNS
    # =========================================================================
    xls = pd.ExcelFile(input_xlsx_path)
    sheet_name = xls.sheet_names[0]  # Targets the first active sheet
    df = pd.read_excel(input_xlsx_path, sheet_name=sheet_name)
    
    # Strip whitespace from column headers to prevent key matching misses
    df.columns = df.columns.str.strip()
    
    # Strict master mapping tables per pillar (+ is Bad/Tight/Headwind, - is Good/Loose/Tailwind)
    pillar_mappings = {
        'Liquidity_Confirmed_Regime': {
             3.0: 'Severe Systemic Crunch', 2.0: 'Significant Deficit', 1.0: 'Mild Deficit',
             0.0: 'Neutral Balanced Anchor', -1.0: 'Mild Surplus', -2.0: 'Significant Surplus', -3.0: 'Hyper-Liquidity Flood'
        },
        'Rates_Regime_Score': {
             3.0: 'Emergency Policy Shock', 2.0: 'Aggressive Hiking Cycle', 1.0: 'Orderly Tightening',
             0.0: 'Policy Rate Fair Value', -1.0: 'Gradual Accommodative Shift', -2.0: 'Aggressive Easing Cycle', -3.0: 'Maximum Policy Stimulus'
        },
        'Confirmed_Inflation_Regime': {
             3.0: 'Runaway Price Shock', 2.0: 'Accelerating Inflation', 1.0: 'Mildly Elevated Inflation',
             0.0: 'Price Stability Equilibrium', -1.0: 'Mild Disinflation', -2.0: 'Deep Deflationary Risk', -3.0: 'Severe Deflationary Spiral'
        },
        'Growth_Confirmed_Regime': {
             3.0: 'Severe Contraction Shock', 2.0: 'Significant Slowdown', 1.0: 'Below-Trend Moderation',
             0.0: 'Trend Growth Equilibrium', -1.0: 'Above-Trend Acceleration', -2.0: 'Robust Economic Boom', -3.0: 'Extreme Macro Expansion'
        },
        'Confirmed External Stress Regime': {
             3.0: 'Extreme External Crisis', 2.0: 'Significant Headwinds', 1.0: 'Mild Capital Tightening',
             0.0: 'Stable External Corridor', -1.0: 'Favorable Cross-Border Inflows', -2.0: 'Strong Capital Saturation', -3.0: 'Extreme Capital Surge'
        }
    }
    
    # Process each pillar and insert the description column immediately next to it
    for col_name, mapping_dict in pillar_mappings.items():
        if col_name in df.columns:
            def map_value(val):
                if pd.isna(val): 
                    return 'Data Unavailable'
                try: 
                    return mapping_dict.get(float(val), f'Unknown Regime ({val})')
                except ValueError: 
                    return f'Unknown Regime ({val})'
            
            qualitative_labels = df[col_name].apply(map_value)
            current_idx = df.columns.get_loc(col_name)
            
            # Dynamic insert at (current_idx + 1) places the label column directly to the right
            df.insert(loc=current_idx + 1, column=col_name + '_Label', value=qualitative_labels)

    # Save to disk as a temporary workbook to hand over to the openpyxl layout engine
    df.to_excel(output_xlsx_path, index=False)

    # =========================================================================
    # STEP 2: OPENPYXL VISUAL BEAUTIFICATION PLUMBING
    # =========================================================================
    wb = openpyxl.load_workbook(output_xlsx_path)
    ws = wb.active
    
    # Enforce grid layout line visibility across Excel versions
    ws.views.sheetView[0].showGridLines = True 

    # Typography & Fill Selections (Executive Classic Slate Navy Theme)
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Navy Blue
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    label_font = Font(name="Segoe UI", size=10, italic=True, color="404040") # Grayed italic text for labels
    num_font = Font(name="Segoe UI", size=10, bold=True, color="000000")    # Crisp bold font for values
    date_font = Font(name="Segoe UI", size=10, color="000000")
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    medium_bottom = Side(border_style="medium", color="1F497D")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Dynamic background alerts mapping (Sophisticated pastel tints to reduce visual noise)
    alert_fills = {
         3.0: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"), # Soft Crimson (Severe Stress)
         2.0: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), # Muted Amber (Significant Stress)
         1.0: PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid"), # Light Warm Tint (Mild Stress)
         0.0: PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"), # Crisp Light Gray (Neutral Balance)
        -1.0: PatternFill(start_color="F2F9EE", end_color="F2F9EE", fill_type="solid"), # Gentle Green (Mild Tailwind)
        -2.0: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # Soft Meadow Green (Significant Tailwind)
        -3.0: PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"), # Stronger Pastel Green (Max Easing/Boom)
    }
    
    # Format Header Row
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=medium_bottom)
        
    # Format Data Rows
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        
        # Format Date tracking column (Col 1 / Column A)
        date_cell = ws.cell(row=row, column=1)
        date_cell.font = date_font
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.border = border_all
        
        # Format metrics and labels
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            header_name = ws.cell(row=1, column=col).value
            cell.border = border_all
            
            # Process text label columns vs numeric metric columns
            if "_Label" in header_name:
                cell.font = label_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Look at the previous column (col-1) to pull the basis value for heatmap tinting
                assoc_score = ws.cell(row=row, column=col-1).value
            else:
                cell.font = num_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '#,##0'
                assoc_score = cell.value
                
            # Inject soft conditional background fills contextually
            if assoc_score is not None and not isinstance(assoc_score, str):
                try:
                    f_score = float(assoc_score)
                    if f_score in alert_fills:
                        cell.fill = alert_fills[f_score]
                except ValueError:
                    pass

    # Clean auto-fit calculation window for column spacing
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)
        
    wb.save(output_xlsx_path)
    print(f"Workbook successfully completed and saved to: {output_xlsx_path}")

# =========================================================================
# PRODUCTION PIPELINE RUNNER
# =========================================================================
if __name__ == "__main__":
    input_file = "Macro Pillars Attempt 2.xlsx"
    output_file = "Macro_Regimes_Formatted_Final.xlsx"
    
    generate_styled_macro_regimes(input_file, output_file)